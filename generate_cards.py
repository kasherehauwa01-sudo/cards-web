#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Консольный генератор карточек со штрихкодами EAN-13 и ФИО.

Скрипт читает данные из Excel (.xlsx), применяет параметры из config.json (если есть
рядом с Excel), строит карточки с помощью Pillow + python-barcode и формирует PDF на
основе ReportLab. Дополнительно можно сохранить отдельные карточки в PNG.

Пример запуска в Windows:
    python generate_cards.py -i D:\\path\\to\\file.xlsx

Важно (Windows): если при копировании файла из diff/почты в начале строк появляются
символы «+» и возникает SyntaxError на первой строке, скачайте исходник целиком (raw)
или откройте файл в редакторе и удалите лишние «+» перед строками.
"""
from __future__ import annotations

import argparse
import os
import json
import logging
import re
import subprocess
import sys
from dataclasses import dataclass, replace
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

from barcode import EAN13
from barcode.writer import ImageWriter
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas

# Константы конвертации и вывода
MM_IN_INCH = 25.4
PT_IN_INCH = 72
DPI = 300
A4_WIDTH_MM = 210
A4_HEIGHT_MM = 297


@dataclass
class Config:
    """Параметры верстки карточек и листа A4."""

    card_width_mm: float = 35
    card_height_mm: float = 15
    card_border_thickness_mm: float = 0.2

    font_path: str = "ArialNarrow.ttf"
    font_size_pt: int = 7
    text_orientation: str = "horizontal"  # horizontal | vertical
    text_top_offset_mm: float = 2

    # Ширина штрихкода в миллиметрах. Высота масштабируется пропорционально
    # фактическому изображению EAN-13, чтобы не искажать штрихкод.
    barcode_width_mm: float = 26.0

    # Коэффициент изменения высоты штрихкода в процентах (ширина неизменна).
    # Например, 120 увеличит высоту на 20 %, 80 уменьшит на 20 %.
    barcode_height_scale_percent: float = 100.0

    # Вертикальный отступ штрихкода в процентах от высоты карточки.
    # Если None — штрихкод центрируется по высоте (с учётом текста сверху).
    barcode_top_offset_percent: Optional[float] = None

    # Смещение штрихкода от правого края карточки в миллиметрах
    barcode_right_offset_mm: float = 6.0

    cards_per_row: int = 5
    gap_mm: float = 1
    top_margin_mm: float = 12

    export_individual_cards: bool = True


DEFAULT_CONFIG = Config()

logger = logging.getLogger("cards")


# ===================== Утилиты конвертации =====================
def mm_to_points(mm_value: float) -> float:
    """Перевод миллиметров в поинты (ReportLab)."""

    return mm_value / MM_IN_INCH * PT_IN_INCH


def mm_to_px(mm_value: float, dpi: int = DPI) -> int:
    """Перевод миллиметров в пиксели для Pillow."""

    return int(round(mm_value / MM_IN_INCH * dpi))


# ===================== Работа с пользователем =====================
def select_input_folder() -> Path:
    """Запрос папки у пользователя до тех пор, пока она не будет существовать."""

    while True:
        user_input = input("Укажите путь к папке с Excel-файлом: ").strip() or "."
        folder = Path(user_input).expanduser().resolve()
        if folder.is_dir():
            return folder
        print("Папка не найдена, попробуйте снова.")


def select_xlsx_file(folder: Path) -> Path:
    """Показать список .xlsx файлов и дать выбрать по номеру или имени."""

    files = sorted(folder.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError("В выбранной папке нет файлов .xlsx")

    print("Найденные Excel-файлы:")
    for idx, path in enumerate(files, start=1):
        print(f"  {idx}. {path.name}")

    while True:
        choice = input("Выберите файл по номеру или имени: ").strip()
        if choice.isdigit():
            pos = int(choice)
            if 1 <= pos <= len(files):
                return files[pos - 1]
        else:
            for path in files:
                if path.name == choice:
                    return path
        print("Неверный выбор, попробуйте снова.")


# ===================== Загрузка и проверка конфигурации =====================
def load_config(base_folder: Path) -> Config:
    """Прочитать config.json рядом с Excel и наложить на значения по умолчанию."""

    config_path = base_folder / "config.json"
    config = replace(DEFAULT_CONFIG)
    if not config_path.exists():
        return config

    try:
        with config_path.open("r", encoding="utf-8") as fh:
            data = json.load(fh)
    except json.JSONDecodeError as exc:  # noqa: BLE001
        # Даём человеку понятное сообщение с указанием строки/столбца.
        msg = (
            f"config.json: ошибка синтаксиса JSON (строка {exc.lineno}, "
            f"столбец {exc.colno}): {exc.msg}"
        )
        logger.error(msg)
        raise ValueError(msg) from exc

    for field_name in config.__dataclass_fields__:
        if field_name in data:
            setattr(config, field_name, data[field_name])

    return validate_config(config)


def validate_config(config: Config) -> Config:
    """Проверить конфигурацию, выбрасывая ошибки и предупреждая о проблемах."""

    errors: List[str] = []
    warnings: List[str] = []

    def ensure_positive(name: str, value: Optional[float], allow_zero: bool = False):
        if value is None:
            return
        if (not allow_zero and value <= 0) or (allow_zero and value < 0):
            errors.append(f"Параметр {name} должен быть положительным, получено {value}")

    ensure_positive("card_width_mm", config.card_width_mm)
    ensure_positive("card_height_mm", config.card_height_mm)
    ensure_positive("card_border_thickness_mm", config.card_border_thickness_mm, allow_zero=True)
    ensure_positive("font_size_pt", config.font_size_pt)
    ensure_positive("text_top_offset_mm", config.text_top_offset_mm, allow_zero=True)
    ensure_positive("gap_mm", config.gap_mm, allow_zero=True)
    ensure_positive("top_margin_mm", config.top_margin_mm, allow_zero=True)

    def ensure_percent(name: str, value: Optional[float]):
        if value is None:
            return
        if value <= 0 or value > 100:
            errors.append(f"Параметр {name} должен быть в диапазоне (0..100], получено {value}")

    ensure_positive("barcode_width_mm", config.barcode_width_mm)
    ensure_percent("barcode_height_scale_percent", config.barcode_height_scale_percent)
    ensure_positive("barcode_right_offset_mm", config.barcode_right_offset_mm, allow_zero=True)
    if config.barcode_top_offset_percent is not None:
        ensure_percent("barcode_top_offset_percent", config.barcode_top_offset_percent)

    if config.cards_per_row < 1:
        errors.append("cards_per_row должен быть не меньше 1")

    orientation = (config.text_orientation or "").lower()
    if orientation not in {"horizontal", "vertical"}:
        warnings.append(
            f"Неизвестная ориентация текста '{config.text_orientation}', используется horizontal"
        )
        config.text_orientation = "horizontal"
    else:
        config.text_orientation = orientation

    barcode_width_mm = config.barcode_width_mm
    if (barcode_width_mm + config.barcode_right_offset_mm) > config.card_width_mm:
        warnings.append("Штрихкод может не поместиться по ширине карточки, скорректируйте ширину/отступ")

    row_width = config.cards_per_row * config.card_width_mm + (config.cards_per_row - 1) * config.gap_mm
    if row_width > A4_WIDTH_MM:
        warnings.append("Карточки в ряд выходят за пределы A4, будет перенос на новую строку")

    if warnings:
        for msg in warnings:
            logger.warning(msg)
    if errors:
        for msg in errors:
            logger.error(msg)
        raise ValueError("Ошибка в config.json, см. лог")

    return config


# ===================== Парсинг данных =====================
def parse_fio(raw: str) -> str:
    """Привести ФИО к формату "Фамилия И.О."."""

    if raw is None:
        raise ValueError("Пустое значение ФИО")

    cleaned = " ".join(str(raw).strip().split())
    if not cleaned:
        raise ValueError("Пустое значение ФИО")

    tokens = re.findall(r"[A-Za-zА-Яа-яЁё\.-]+", cleaned)
    if not tokens:
        raise ValueError("Не удалось распознать ФИО")

    surname = tokens[0].replace("ё", "Ё").upper()

    if any("." in t for t in tokens[1:]):
        initials = " ".join(t.upper() for t in tokens[1:])
        return f"{surname} {initials}".strip()

    initials_parts = [t for t in tokens[1:] if t.strip(".-")]
    if not initials_parts:
        return surname

    initials = ".".join(part[0].upper() for part in initials_parts[:2]) + "."
    return f"{surname} {initials}".strip()


def compute_ean13_check_digit(data12: str) -> str:
    """Вычислить контрольную цифру для 12-значной строки."""

    if len(data12) != 12 or not data12.isdigit():
        raise ValueError("Для контрольной цифры нужна 12-значная строка")

    digits = [int(ch) for ch in data12]
    checksum = (10 - ((sum(digits[::2]) + sum(d * 3 for d in digits[1::2])) % 10)) % 10
    return str(checksum)


def normalize_barcode(value, row_index: int) -> str:
    """Преобразовать значение из Excel к валидному EAN-13."""

    if value is None:
        raise ValueError(f"Пустой штрихкод в строке {row_index}")

    if isinstance(value, (int, float)):
        raw = format(value, ".0f")
    else:
        raw = str(value).strip()

    digits_only = re.sub(r"\D", "", raw)

    if len(digits_only) == 12:
        digits_only += compute_ean13_check_digit(digits_only)
    elif len(digits_only) == 13:
        expected = compute_ean13_check_digit(digits_only[:12])
        if expected != digits_only[-1]:
            raise ValueError(
                f"Неверная контрольная цифра в строке {row_index}: ожидается {expected}"
            )
    else:
        raise ValueError(
            f"Некорректное количество цифр в строке {row_index}: найдено {len(digits_only)}"
        )

    return digits_only


def check_duplicates(rows: Sequence[Tuple[int, str]]) -> Dict[str, List[int]]:
    """Найти дублирующиеся штрихкоды и вернуть словарь barcode -> номера строк."""

    seen: Dict[str, List[int]] = {}
    for row_idx, barcode_value in rows:
        seen.setdefault(barcode_value, []).append(row_idx)
    return {code: idxs for code, idxs in seen.items() if len(idxs) > 1}


# ===================== Рендер карточки =====================
def resolve_font_path(font_path: Path, assets_dir: Path) -> Optional[Path]:
    """Подобрать доступный путь к шрифту.

    Порядок проверки:
    1. Явно указанный путь (абсолютный или относительный к assets_dir).
    2. Папка Fonts в Windows (если переменная окружения WINDIR задана).
    3. Стандартный DejaVuSans из поставки Pillow (если доступен).
    """

    candidates: List[Path] = []

    if font_path.is_absolute():
        candidates.append(font_path)
    else:
        candidates.append(assets_dir / font_path)
        windir = os.environ.get("WINDIR")
        if windir:
            win_fonts = Path(windir) / "Fonts"
            candidates.append(win_fonts / font_path.name)
            # Дополнительные варианты для Arial Narrow в Windows
            if font_path.stem.lower() in {"arialnarrow", "arial narrow"}:
                candidates.append(win_fonts / "ARIALN.TTF")
                candidates.append(win_fonts / "arialn.ttf")
                candidates.append(win_fonts / "arial.ttf")

    # DejaVu из Pillow как резерв
    pil_font_path = Path(ImageFont.__file__).with_name("DejaVuSans.ttf")
    candidates.append(pil_font_path)

    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def load_font(font_path: Path, font_size: int, assets_dir: Path) -> ImageFont.FreeTypeFont:
    """Попробовать загрузить шрифт, при ошибке — вернуть встроенный."""

    resolved = resolve_font_path(font_path, assets_dir)
    if resolved:
        try:
            return ImageFont.truetype(str(resolved), font_size)
        except Exception as exc:  # noqa: BLE001
            logger.warning("Не удалось загрузить шрифт %s (%s), fallback", resolved, exc)

    # Попытка загрузить DejaVuSans из ресурсов Pillow по имени
    try:
        return ImageFont.truetype("DejaVuSans.ttf", font_size)
    except Exception as exc:  # noqa: BLE001
        logger.warning("Не удалось загрузить DejaVuSans.ttf из Pillow (%s)", exc)

    logger.warning(
        "Не удалось найти шрифт %s, используется встроенный Pillow (может не поддерживать кириллицу)",
        font_path,
    )
    return ImageFont.load_default()


def measure_text(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.FreeTypeFont) -> Tuple[int, int]:
    """Совместимое измерение текста для Pillow >=10 (textbbox) и старых версий (textsize)."""

    if hasattr(draw, "textbbox"):
        bbox = draw.textbbox((0, 0), text, font=font)
        return bbox[2] - bbox[0], bbox[3] - bbox[1]
    return draw.textsize(text, font=font)


def make_barcode_image(code: str, target_width_px: int, height_scale_percent: float) -> Image.Image:
    """Создать изображение штрихкода и масштабировать его по ширине.

    Высота дополнительно умножается на коэффициент height_scale_percent, чтобы
    можно было сделать штрихкод выше/ниже без изменения ширины (требование
    макета сохранить пропорции по ширине).
    """

    # «write_text=False» отключает подпись цифр под штрихкодом — по требованию макета.
    writer = ImageWriter()
    ean = EAN13(code[:-1], writer=writer)  # библиотека сама добавит контрольную цифру

    buffer = BytesIO()
    ean.write(buffer, options={"write_text": False})
    buffer.seek(0)
    barcode_img = Image.open(buffer).convert("RGB")

    # Масштабируем пропорционально по ширине (высота тянется автоматически),
    # чтобы не было непропорциональных искажений штрихкода. Затем умножаем
    # высоту на коэффициент из конфигурации, не трогая ширину.
    if target_width_px <= 0:
        raise ValueError("Ширина штрихкода должна быть положительной")

    width, height = barcode_img.size
    scale = target_width_px / float(width)
    height_scale = max(height_scale_percent, 1e-3) / 100.0
    target_height_px = max(1, int(round(height * scale * height_scale)))
    return barcode_img.resize((target_width_px, target_height_px), resample=Image.LANCZOS)


def draw_card(config: Config, fio: str, barcode_value: str, assets_dir: Path) -> Image.Image:
    """Собрать карточку в Pillow для дальнейшего экспорта."""

    card_w_px = mm_to_px(config.card_width_mm)
    card_h_px = mm_to_px(config.card_height_mm)
    border_px = max(1, mm_to_px(config.card_border_thickness_mm))

    card = Image.new("RGB", (card_w_px, card_h_px), "white")
    draw = ImageDraw.Draw(card)

    # Рисуем рамку
    draw.rectangle([(0, 0), (card_w_px - 1, card_h_px - 1)], outline="black", width=border_px)

    # Текст ФИО
    font_path = Path(config.font_path)
    font = load_font(font_path, int(config.font_size_pt), assets_dir)

    text_y = mm_to_px(config.text_top_offset_mm)
    text_w, text_h = measure_text(draw, fio, font)
    text_x = max((card_w_px - text_w) // 2, border_px)

    if config.text_orientation == "horizontal":
        draw.text((text_x, text_y), fio, fill="black", font=font)
    else:
        # Создаем отдельное изображение для вращения текста
        text_layer = Image.new("RGBA", (text_w, text_h), (255, 255, 255, 0))
        text_draw = ImageDraw.Draw(text_layer)
        text_draw.text((0, 0), fio, fill="black", font=font)
        rotated = text_layer.rotate(90, expand=True)
        rx, ry = rotated.size
        card.paste(rotated, ((card_w_px - rx) // 2, text_y), rotated)

    # Штрихкод
    barcode_target_width_px = mm_to_px(config.barcode_width_mm)
    barcode_img = make_barcode_image(
        barcode_value,
        barcode_target_width_px,
        config.barcode_height_scale_percent,
    )
    barcode_w, barcode_h = barcode_img.size

    barcode_x = card_w_px - mm_to_px(config.barcode_right_offset_mm) - barcode_w
    if config.barcode_top_offset_percent is None:
        barcode_y = max((card_h_px - barcode_h) // 2, text_y + text_h + 2)
    else:
        barcode_y = int(card_h_px * (config.barcode_top_offset_percent / 100.0))

    barcode_x = max(border_px, barcode_x)
    barcode_y = max(border_px, barcode_y)

    card.paste(barcode_img, (barcode_x, barcode_y))
    return card


# ===================== Компоновка на листе A4 =====================
def layout_cards_on_a4(cards: Sequence[Image.Image], config: Config, output_pdf: Path):
    """Разложить карточки по листам A4 и сохранить PDF."""

    page_w_pt = mm_to_points(A4_WIDTH_MM)
    page_h_pt = mm_to_points(A4_HEIGHT_MM)
    card_w_pt = mm_to_points(config.card_width_mm)
    card_h_pt = mm_to_points(config.card_height_mm)
    gap_pt = mm_to_points(config.gap_mm)
    top_margin_pt = mm_to_points(config.top_margin_mm)

    pdf = canvas.Canvas(str(output_pdf), pagesize=(page_w_pt, page_h_pt))

    col = 0
    x = 0.0
    y = page_h_pt - top_margin_pt - card_h_pt

    for card_img in cards:
        if col >= config.cards_per_row:
            col = 0
            x = 0.0
            y -= card_h_pt + gap_pt

        if y < 0:
            pdf.showPage()
            col = 0
            x = 0.0
            y = page_h_pt - top_margin_pt - card_h_pt

        img_reader = ImageReader(card_img)
        pdf.drawImage(img_reader, x, y, width=card_w_pt, height=card_h_pt, preserveAspectRatio=False)

        col += 1
        x += card_w_pt + gap_pt

    pdf.save()


# ===================== Открытие PDF =====================
def open_pdf_file(pdf_path: Path):
    """Открыть сгенерированный PDF штатным просмотрщиком ОС."""

    if not pdf_path.exists():
        logger.warning("PDF %s не найден для открытия", pdf_path)
        return

    try:
        # На Windows сначала пробуем os.startfile, а если система блокирует
        # (групповые политики/ассоциации), используем запасной вариант через cmd /c start.
        if sys.platform.startswith("win"):
            try:
                os.startfile(str(pdf_path))  # type: ignore[attr-defined]
            except OSError:
                subprocess.Popen(
                    ["cmd", "/c", "start", "", str(pdf_path)],
                    shell=True,
                )
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(pdf_path)])
        else:
            subprocess.Popen(["xdg-open", str(pdf_path)])
        logger.info("Открыт PDF: %s", pdf_path)
    except Exception as exc:  # noqa: BLE001
        logger.warning("Не удалось автоматически открыть PDF: %s", exc)


# ===================== Экспорт PNG =====================
def export_individual_cards(cards: Sequence[Tuple[str, str, Image.Image]], output_dir: Path):
    """Сохранить каждую карточку в PNG с именем по ФИО или штрихкоду."""

    output_dir.mkdir(parents=True, exist_ok=True)
    for fio, barcode_value, card_img in cards:
        name_part = fio if fio else barcode_value
        safe_name = re.sub(r"[^\w\-\.]+", "_", name_part)[:100]
        file_path = output_dir / f"{safe_name or barcode_value}.png"
        card_img.save(file_path, dpi=(DPI, DPI))


# ===================== Чтение Excel =====================
def read_excel_rows(xlsx_path: Path) -> List[Tuple[int, str, str]]:
    """Считать строки Excel и вернуть список (row_idx, fio, barcode)."""

    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    sheet = wb.active

    entries: List[Tuple[int, str, str]] = []
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        # Пропускаем первую строку как заголовок
        if idx == 1:
            continue

        values = list(row)
        # Пропускаем полностью пустые строки (включая ячейки с пустыми строками/пробелами)
        if not values or all((val is None) or (isinstance(val, str) and not val.strip()) for val in values):
            continue

        fio_raw = values[0] if len(values) > 0 else None
        barcode_raw = values[1] if len(values) > 1 else None

        # Обрабатываем только строки, где в колонке ФИО есть содержимое
        if fio_raw is None or (isinstance(fio_raw, str) and not fio_raw.strip()):
            continue

        # Предварительно очищаем ячейку штрихкода: оставляем только цифры, иначе считаем пустой
        if barcode_raw is not None:
            digits = re.sub(r"\D", "", str(barcode_raw).strip())
            if len(digits) not in {12, 13}:
                logger.warning("Строка %s: штрихкод очищен как пустой (длина %s)", idx, len(digits))
                barcode_raw = None
            else:
                barcode_raw = digits
        else:
            logger.warning("Строка %s: штрихкод пустой, строка будет пропущена", idx)

        if barcode_raw is None:
            continue

        try:
            fio = parse_fio(fio_raw)
            barcode = normalize_barcode(barcode_raw, idx)
        except Exception as exc:  # noqa: BLE001
            raise ValueError(f"Строка {idx}: {exc}") from exc

        entries.append((idx, fio, barcode))

    if not entries:
        raise ValueError("В Excel нет данных для обработки")

    return entries


# ===================== Логирование =====================
def setup_logging(log_path: Path):
    """Настроить логирование в файл и консоль.

    Создает родительские каталоги при необходимости, записывает логи в файл
    `cards.log` в папке Excel и дублирует сообщения в stdout.
    """

    logger.setLevel(logging.INFO)
    formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")

    log_path.parent.mkdir(parents=True, exist_ok=True)

    file_handler = logging.FileHandler(log_path, encoding="utf-8")
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(formatter)

    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)

    logger.handlers.clear()
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    logger.propagate = False


# ===================== Помощники =====================
def prompt_yes_no(message: str) -> bool:
    """Запросить подтверждение у пользователя (Y/N)."""

    while True:
        choice = input(f"{message} [Y/N]: ").strip().lower()
        if choice in {"y", "yes", "д", "да"}:
            return True
        if choice in {"n", "no", "н", "нет"}:
            return False
        print("Введите Y или N.")


def resolve_output_pdf(excel_path: Path) -> Path:
    """Определить имя выходного PDF с учетом существующих файлов."""

    base = excel_path.with_suffix("")
    candidate = base.with_suffix(".pdf")
    if not candidate.exists():
        return candidate

    logger.info("Файл %s уже существует", candidate.name)
    overwrite = prompt_yes_no("Перезаписать существующий PDF?")
    if overwrite:
        return candidate

    idx = 2
    while True:
        new_candidate = base.with_name(f"{base.name}_v{idx}").with_suffix(".pdf")
        if not new_candidate.exists():
            return new_candidate
        idx += 1


# ===================== Основной сценарий =====================
def process_file(xlsx_path: Path):
    """Полный цикл обработки одного Excel."""

    folder = xlsx_path.parent
    setup_logging(folder / "cards.log")
    logger.info("Логирование включено: %s", folder / "cards.log")
    logger.info("Входной файл: %s", xlsx_path.name)

    config = load_config(folder)
    logger.info("Конфигурация: %s", config)

    entries = read_excel_rows(xlsx_path)
    logger.info("Прочитано строк: %s", len(entries))

    duplicates = check_duplicates([(row_idx, barcode) for row_idx, _, barcode in entries])
    if duplicates:
        dup_list = ", ".join(f"{code} (строки {rows})" for code, rows in duplicates.items())
        logger.warning("Найдены дубли: %s", dup_list)
        if not prompt_yes_no("Продолжить обработку несмотря на дубли?"):
            logger.info("Работа остановлена пользователем из-за дублей")
            return

    cards: List[Tuple[str, str, Image.Image]] = []
    for _, fio, barcode in entries:
        card_img = draw_card(config, fio, barcode, folder)
        cards.append((fio, barcode, card_img))

    output_pdf = resolve_output_pdf(xlsx_path)
    layout_cards_on_a4([card for _, _, card in cards], config, output_pdf)
    logger.info("PDF сохранен: %s", output_pdf)

    # Открываем готовый PDF, чтобы пользователь сразу увидел результат
    open_pdf_file(output_pdf)

    if config.export_individual_cards:
        export_dir = output_pdf.with_name("cards")
        export_individual_cards(cards, export_dir)
        logger.info("Экспорт отдельных карточек: %s", export_dir)


def build_arg_parser() -> argparse.ArgumentParser:
    """Сформировать парсер аргументов командной строки."""

    parser = argparse.ArgumentParser(
        description="Генерация карточек со штрихкодами EAN-13 из Excel",
    )
    parser.add_argument(
        "--input",
        "-i",
        type=Path,
        help="Путь к Excel-файлу (.xlsx). Если не указан, будет интерактивный выбор.",
    )
    return parser


def main(argv: Optional[Sequence[str]] = None):
    """Точка входа CLI."""

    parser = build_arg_parser()
    args = parser.parse_args(argv)

    if args.input:
        xlsx_path = args.input.expanduser().resolve()
        if not xlsx_path.exists():
            print(f"Файл {xlsx_path} не найден")
            sys.exit(1)
    else:
        folder = Path.cwd()
        xlsx_path = select_xlsx_file(folder)

    try:
        process_file(xlsx_path)
    except Exception as exc:  # noqa: BLE001
        logger.exception("Критическая ошибка: %s", exc)
        print(f"Ошибка: {exc}")
        sys.exit(1)

def process_file_web(xlsx_path: Path) -> Path:
    """Версия для веба: без input() и без открытия PDF."""
    folder = xlsx_path.parent
    setup_logging(folder / "cards.log")

    config = load_config(folder)
    entries = read_excel_rows(xlsx_path)

    duplicates = check_duplicates([(row_idx, barcode) for row_idx, _, barcode in entries])
    if duplicates:
        dup_list = ", ".join(f"{code} (строки {rows})" for code, rows in duplicates.items())
        logger.warning("Найдены дубли: %s", dup_list)

    cards: List[Tuple[str, str, Image.Image]] = []
    for _, fio, barcode in entries:
        card_img = draw_card(config, fio, barcode, folder)
        cards.append((fio, barcode, card_img))

    output_pdf = xlsx_path.with_suffix(".pdf")
    layout_cards_on_a4([card for _, _, card in cards], config, output_pdf)

    return output_pdf


if __name__ == "__main__":
    main()
